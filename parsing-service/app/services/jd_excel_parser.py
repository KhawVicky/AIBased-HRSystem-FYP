"""Extracts job details from an Excel worksheet."""

import re
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from app.utils.text_cleaner import (
    clean_text,
    normalized_heading,
    strip_section_number,
    strip_value_prefix,
)


JOB_TITLE_LABELS = ("JOB TITLE", "POSITION")
DEPARTMENT_LABELS = ("DEPARTMENT",)
SALARY_LABELS = (
    "SALARY",
    "SALARY RANGE",
    "BASIC SALARY",
    "MONTHLY SALARY",
    "REMUNERATION",
)
QUALIFICATION_HEADINGS = ("qualification", "qualifications")
DESCRIPTION_HEADINGS = ("job purpose",)
RESPONSIBILITY_HEADINGS = (
    "responsibilities as follows but not limited to",
    "job responsibilities",
    "key responsibilities",
    "responsibilities",
    "responsibility",
    "duties",
)
RESPONSIBILITY_STOP_HEADINGS = (
    "qualification",
    "qualifications",
    "requirement",
    "requirements",
    "education",
    "experience",
    "skills",
    "competencies",
    "authority",
    "authorities",
)


class JDParsingError(Exception):
    def __init__(self, code: str, message: str, status_code: int = 400) -> None:
        super().__init__(message)
        self.code = code
        self.message = message
        self.status_code = status_code


def validate_xlsx_filename(filename: str | None) -> None:
    if not filename:
        raise JDParsingError("FILE_REQUIRED", "An Excel file is required.")
    if Path(filename).suffix.lower() != ".xlsx":
        raise JDParsingError("INVALID_FILE_TYPE", "Only .xlsx files are supported.")


def load_excel_workbook(content: bytes):
    if not content:
        raise JDParsingError("EMPTY_FILE", "The uploaded Excel file is empty.")
    # Read-only mode keeps large workbooks light in memory.
    try:
        return load_workbook(BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:
        raise JDParsingError(
            "INVALID_EXCEL", "The Excel file could not be opened."
        ) from exc


def _worksheet_rows(worksheet: Worksheet) -> list[list[str]]:
    # Remove empty cells while keeping the visible row order.
    rows: list[list[str]] = []
    for excel_row in worksheet.iter_rows(values_only=True):
        cells = [clean_text(value) for value in excel_row]
        while cells and not cells[-1]:
            cells.pop()
        if any(cells):
            rows.append(cells)
    return rows


def _compact_cells(row: list[str]) -> list[str]:
    return [cell for cell in row if cell]


def _sentence_case_if_uppercase(value: str) -> str:
    # Keep normal mixed-case text exactly as written in Excel.
    value = clean_text(value)
    if not value or not value.isupper():
        return value
    return value[:1].upper() + value[1:].lower()


def _label_value(rows: list[list[str]], labels: tuple[str, ...]) -> str:
    # Support values in the same cell or in a later cell.
    for row in rows:
        cells = _compact_cells(row)
        for index, cell in enumerate(cells):
            for label in labels:
                match = re.match(
                    rf"^\s*{re.escape(label)}(?:\s*\([^)]*\))?\s*(?::\s*(.*))?$",
                    cell,
                    flags=re.IGNORECASE,
                )
                if not match:
                    continue
                inline_value = strip_value_prefix(match.group(1) or "")
                if inline_value:
                    return inline_value
                for later_cell in cells[index + 1 :]:
                    later_value = strip_value_prefix(later_cell)
                    if later_value:
                        return later_value
    return ""


def _heading_inline_value(cell: str, headings: tuple[str, ...]) -> str | None:
    normalized = normalized_heading(cell)
    heading_cell = strip_section_number(cell)
    for heading in headings:
        if normalized == heading:
            return ""
        match = re.match(
            rf"^\s*{re.escape(heading)}\s*:\s*(.+)$",
            heading_cell,
            flags=re.IGNORECASE,
        )
        if match:
            return clean_text(match.group(1))
    return None


def _responsibility_heading_value(cell: str) -> str | None:
    normalized = normalized_heading(cell)
    heading_cell = strip_section_number(cell)
    for heading in RESPONSIBILITY_HEADINGS:
        if normalized == heading:
            return ""
        match = re.match(
            rf"^\s*{re.escape(heading)}\s*(?::\s*(.*))?$",
            heading_cell,
            flags=re.IGNORECASE,
        )
        if match:
            return clean_text(match.group(1) or "")
        if normalized.startswith(f"{heading} "):
            return ""
    return None


def _is_stop_heading(row_text: str, headings: tuple[str, ...]) -> bool:
    normalized = normalized_heading(row_text)
    return any(
        normalized == heading or normalized.startswith(f"{heading}:")
        for heading in headings
    )


def _strip_list_number(value: str) -> tuple[str, bool]:
    value = clean_text(value)
    if re.fullmatch(r"\d+[.),]?", value):
        return "", True
    match = re.match(r"^\s*(?:[-*•]\s*)?(\d+)\s*[.),|:-]\s*(.+)$", value)
    if match:
        return clean_text(match.group(2)), True
    return re.sub(r"^\s*[-*•]\s*", "", value).strip(), False


def _row_item(cells: list[str]) -> tuple[str, bool]:
    compact = _compact_cells(cells)
    if not compact:
        return "", False
    numbered = False
    if re.fullmatch(r"\d+[.),]?", compact[0]):
        compact = compact[1:]
        numbered = True
    value, inline_numbered = _strip_list_number(" ".join(compact))
    return value, numbered or inline_numbered


def _append_item(items: list[str], value: str, numbered: bool) -> None:
    if not value:
        return
    # Join wrapped rows only when they look like sentence continuations.
    if (
        items
        and not numbered
        and value[:1].islower()
        and not items[-1].endswith((".", ";", ":", "?", "!"))
    ):
        items[-1] = f"{items[-1]} {value}"
    else:
        items.append(value)


def _extract_sections(rows: list[list[str]]) -> tuple[str, list[str], list[str]]:
    # Move through the sheet until a known heading changes the active section.
    description_rows: list[str] = []
    qualifications: list[str] = []
    responsibilities: list[str] = []
    section: str | None = None

    for row in rows:
        cells = _compact_cells(row)
        if not cells:
            continue

        heading_index = None
        inline_value = ""
        next_section = None
        for index, cell in enumerate(cells):
            description_value = _heading_inline_value(cell, DESCRIPTION_HEADINGS)
            if description_value is not None:
                heading_index = index
                inline_value = description_value
                next_section = "description"
                break
            qualification_value = _heading_inline_value(cell, QUALIFICATION_HEADINGS)
            if qualification_value is not None:
                heading_index = index
                inline_value = qualification_value
                next_section = "qualifications"
                break
            responsibility_value = _responsibility_heading_value(cell)
            if responsibility_value is not None:
                heading_index = index
                inline_value = responsibility_value
                next_section = "responsibilities"
                break

        if next_section:
            section = next_section
            following = cells[(heading_index or 0) + 1 :]
            if following:
                following[0] = strip_value_prefix(following[0])
            content = " ".join(part for part in [inline_value, *following] if part)
            value, numbered = _strip_list_number(content)
            target = (
                description_rows
                if section == "description"
                else qualifications
                if section == "qualifications"
                else responsibilities
            )
            _append_item(target, value, numbered)
            continue

        row_text = " ".join(cells)
        if section == "description":
            value, numbered = _row_item(cells)
            _append_item(description_rows, value, numbered)
        elif section == "qualifications":
            if _is_stop_heading(row_text, ("job description",)):
                section = None
                continue
            value, numbered = _row_item(cells)
            _append_item(qualifications, value, numbered)
        elif section == "responsibilities":
            if _is_stop_heading(row_text, RESPONSIBILITY_STOP_HEADINGS):
                section = None
                continue
            value, numbered = _row_item(cells)
            _append_item(responsibilities, value, numbered)

    return "\n".join(description_rows), qualifications, responsibilities


def _sheet_summary(worksheet: Worksheet, rows: list[list[str]]) -> dict[str, str]:
    return {
        "sheetName": worksheet.title,
        "jobTitle": _sentence_case_if_uppercase(
            _label_value(rows, JOB_TITLE_LABELS) or worksheet.title
        ),
        "department": _sentence_case_if_uppercase(
            _label_value(rows, DEPARTMENT_LABELS)
        ),
    }


def list_worksheets(content: bytes, filename: str | None) -> dict:
    # Empty worksheets are hidden from the HR selection dialog.
    validate_xlsx_filename(filename)
    workbook = load_excel_workbook(content)
    sheets = []
    try:
        for worksheet in workbook.worksheets:
            rows = _worksheet_rows(worksheet)
            if rows:
                sheets.append(_sheet_summary(worksheet, rows))
    finally:
        workbook.close()

    if not sheets:
        raise JDParsingError(
            "NO_VALID_WORKSHEETS", "The workbook contains no non-empty worksheets."
        )
    return {
        "success": True,
        "fileName": filename,
        "totalSheets": len(sheets),
        "sheets": sheets,
    }


def extract_worksheet(content: bytes, filename: str | None, sheet_name: str) -> dict:
    # Return editable job fields plus the original text for later analysis.
    validate_xlsx_filename(filename)
    if not clean_text(sheet_name):
        raise JDParsingError("SHEET_NAME_REQUIRED", "A worksheet name is required.")

    workbook = load_excel_workbook(content)
    try:
        if sheet_name not in workbook.sheetnames:
            raise JDParsingError(
                "SHEET_NOT_FOUND", "The selected worksheet does not exist.", 404
            )
        worksheet = workbook[sheet_name]
        rows = _worksheet_rows(worksheet)
        if not rows:
            raise JDParsingError(
                "EMPTY_WORKSHEET", "The selected worksheet contains no content.", 422
            )

        job_title = _label_value(rows, JOB_TITLE_LABELS)
        warnings = []
        if not job_title:
            job_title = worksheet.title
            warnings.append(
                "Job title was not found. Sheet name was used as fallback."
            )
        job_title = _sentence_case_if_uppercase(job_title)

        description, qualifications, responsibilities = _extract_sections(rows)
        data = {
            "sheetName": worksheet.title,
            "jobTitle": job_title,
            "department": _sentence_case_if_uppercase(
                _label_value(rows, DEPARTMENT_LABELS)
            ),
            "salary": _label_value(rows, SALARY_LABELS),
            "description": description,
            "qualifications": qualifications,
            "responsibilities": responsibilities,
            "requirements": list(qualifications),
            "rawText": "\n".join(" ".join(_compact_cells(row)) for row in rows),
        }
        return {"success": True, "data": data, "warnings": warnings}
    finally:
        workbook.close()
